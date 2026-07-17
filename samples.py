import os
import re
import statistics
from collections import defaultdict

import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

# ============================================================
# SETTINGS
# ============================================================

DATA_FOLDER = "samples data"
OUTPUT_FILE = "samples.xlsx"
PLOT_FILE = "samples_plot.png"
PLOT_2_FILE = "samples_plot_2.png"

# ============================================================
# STORAGE
# ============================================================

measurements = []
samples = defaultdict(dict)

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def format_sample_name(raw_name):
    """
    Example:
    1Sn-15Au-etch -> 1Sn 15Au etched
    """
    formatted_sample = raw_name.replace("-", " ")

    formatted_sample = re.sub(
        r"\betch\b",
        "etched",
        formatted_sample,
        flags=re.IGNORECASE
    )

    formatted_sample = re.sub(
        r"\bload\b",
        "loaded",
        formatted_sample,
        flags=re.IGNORECASE
    )

    return formatted_sample


def sign_of_number(value):
    return 1 if value > 0 else -1


# ============================================================
# READ FILES
# ============================================================

for filename in os.listdir(DATA_FOLDER):

    filepath = os.path.join(DATA_FOLDER, filename)

    if not os.path.isfile(filepath):
        continue

    # Skip files with extensions
    if "." in filename:
        continue

    # --------------------------------------------------------
    # Extract information from filename
    # --------------------------------------------------------

    parts = filename.split("_")

    if len(parts) < 3:
        print(f"Missing information in filename {filename}")
        break

    raw_sample = parts[0]
    sample = format_sample_name(raw_sample)

    B_field = None
    for part in parts:
        if part in ("B+", "B0", "B-"):
            B_field = part
            break

    if B_field is None:
        continue

    # --------------------------------------------------------
    # Read file
    # --------------------------------------------------------

    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    # --------------------------------------------------------
    # Angle of incidence
    # --------------------------------------------------------

    angle_matches = re.findall(r"angle-of-incidence:\s*([-\d.]+)",text)

    if not angle_matches:
        print(f"No angle found in {filename}")
        break

    angles = {abs(float(angle)) for angle in angle_matches}

    if len(angles) != 1:
        print(f"Inconsistent angle-of-incidence for {filename}")
        break

    angle = angles.pop()

    # --------------------------------------------------------
    # P values
    # --------------------------------------------------------

    p_matches = re.findall(r"\bP\s+(-?\d+(?:\.\d+)?)",text)

    if not p_matches:
        print(f"No P values found in {filename}")
        break

    p_signs = {sign_of_number(float(p_value)) for p_value in p_matches}

    if len(p_signs) != 1:
        print(f"Inconsistent sign of polarization angle in {filename}")
        break

    p_sign = p_signs.pop()

    # --------------------------------------------------------
    # Psi values
    # --------------------------------------------------------

    psi_matches = re.findall(r"Psi:\s*([-\d.]+)",text)

    psi_values = [float(x) for x in psi_matches]

    if not psi_values:
        print(f"No Psi values found in {filename}")
        break

    psi_avg = statistics.mean(psi_values)

    if len(psi_values) > 1:
        psi_std = statistics.stdev(psi_values)
    else:
        psi_std = 0.0

    # --------------------------------------------------------
    # Save measurement
    # --------------------------------------------------------

    measurement = {
        "sample": sample,
        "angle": angle,
        "B": B_field,
        "p_sign": p_sign,
        "psi_avg": psi_avg,
        "psi_std": psi_std,
        "psi_values": psi_values
    }

    measurements.append(measurement)

    samples[sample][B_field] = measurement

# ============================================================
# CREATE EXCEL WORKBOOK
# ============================================================

wb = Workbook()
ws = wb.active
ws.title = "Faraday Analysis"

# ============================================================
# TABLE 1
# ============================================================

current_row = 1

ws.cell(current_row, 1, "Measurements varying sample")
ws.cell(current_row, 1).font = Font(bold=True, size=14)

current_row += 2

header_row = current_row

ws.cell(header_row, 1, "Sample")
ws.cell(header_row, 2, "Angle (°)")
ws.cell(header_row, 3, "Magnetic flux density (B)")
ws.cell(header_row, 4, "P sign")

ws.merge_cells(
    start_row=header_row,
    start_column=5,
    end_row=header_row,
    end_column=7
)

ws.cell(header_row, 5, "Ψ (°)")
for i in range(5):
    ws.cell(header_row, i+1).alignment = Alignment(horizontal="center")

current_row += 1

for m in measurements:

    ws.cell(current_row, 1, m["sample"])
    ws.cell(current_row, 2, m["angle"])
    ws.cell(current_row, 3, m["B"])
    ws.cell(current_row, 4, m["p_sign"])

    ws.cell(current_row, 5, m["psi_avg"])
    ws.cell(current_row, 6, "±")
    ws.cell(current_row, 7, m["psi_std"])

    current_row += 1

# ============================================================
# STORE FARADAY RESULTS FOR PLOTTING
# ============================================================

faraday_rows = []

for sample, data in samples.items():

    if not all(key in data for key in ("B+", "B0", "B-")):
        continue

    psi_plus = data["B+"]["psi_avg"]
    psi_zero = data["B0"]["psi_avg"]
    psi_minus = data["B-"]["psi_avg"]

    std_plus = data["B+"]["psi_std"]
    std_zero = data["B0"]["psi_std"]
    std_minus = data["B-"]["psi_std"]

    p_sign = data["B+"]["p_sign"]

    phi_plus = (psi_plus - psi_zero) * p_sign
    phi_avg = ((psi_plus - psi_minus) / 2) * p_sign
    phi_minus = (psi_zero - psi_minus) * p_sign

    phi_plus_unc = std_plus + std_zero
    phi_avg_unc = std_plus + std_minus
    phi_minus_unc = std_zero + std_minus

    faraday_rows.append(
        {
            "Sample": sample,
            "P sign": p_sign,
            "Phi B+": phi_plus,
            "Phi avg": phi_avg,
            "Phi B-": phi_minus,
            "Unc B+": phi_plus_unc,
            "Unc avg": phi_avg_unc,
            "Unc B-": phi_minus_unc
        }
    )

# ============================================================
# TABLE 2
# ============================================================

current_row += 3

table2_title_row = current_row

ws.cell(
    table2_title_row,
    1,
    "Faraday rotation calculation varying sample"
)
ws.cell(
    table2_title_row,
    1
).font = Font(bold=True, size=14)

current_row += 2

header_row = current_row

ws.cell(header_row, 1, "Sample")
ws.cell(header_row, 2, "P sign")

ws.merge_cells(
    start_row=header_row,
    start_column=3,
    end_row=header_row,
    end_column=5
)

ws.merge_cells(
    start_row=header_row,
    start_column=6,
    end_row=header_row,
    end_column=8
)

ws.merge_cells(
    start_row=header_row,
    start_column=9,
    end_row=header_row,
    end_column=11
)

ws.cell(header_row, 3, "Ψ_B+ - Ψ_B0 (°)")
ws.cell(header_row, 6, "(Ψ_B+ - Ψ_B-)/2 (°)")
ws.cell(header_row, 9, "Ψ_B0 - Ψ_B- (°)")

for col in [3, 6, 9]:
    ws.cell(header_row, col).alignment = Alignment(horizontal="center")

current_row += 1

plot_start_row = current_row

for row in faraday_rows:

    ws.cell(current_row, 1, row["Sample"])
    ws.cell(current_row, 2, row["P sign"])

    ws.cell(current_row, 3, row["Phi B+"])
    ws.cell(current_row, 4, "±")
    ws.cell(current_row, 5, row["Unc B+"])

    ws.cell(current_row, 6, row["Phi avg"])
    ws.cell(current_row, 7, "±")
    ws.cell(current_row, 8, row["Unc avg"])

    ws.cell(current_row, 9, row["Phi B-"])
    ws.cell(current_row, 10, "±")
    ws.cell(current_row, 11, row["Unc B-"])

    current_row += 1

plot_end_row = current_row - 1

# ============================================================
# CREATE PLOT
# ============================================================

plot_data = {}
plot_unc = {}

for row in faraday_rows:
    sample = row["Sample"]
    plot_data[sample] = [
        row["Phi B+"],
        row["Phi avg"],
        row["Phi B-"]
    ]
    plot_unc[sample] = [
    row["Unc B+"],
    row["Unc avg"],
    row["Unc B-"]
    ]

plt.figure(figsize=(8, 6))

x = [0, 1, 2]
x_labels = ["Ψ_B+ - Ψ_B0", "(Ψ_B+ - Ψ_B-)/2", "Ψ_B0 - Ψ_B-"]

for sample in sorted(plot_data.keys()):

    plt.errorbar(
        x,
        plot_data[sample],
        yerr=plot_unc[sample],
        fmt='o-',
        capsize=6,
        linewidth=2,
        markersize=6,
        label=sample
    )

plt.xticks(x, x_labels)

plt.ylabel("ϕ (°)")
plt.xlabel("ϕ calculation method")
plt.title("Faraday rotation (ϕ) for different samples")

plt.grid(True, linestyle='--', alpha=0.6)
plt.legend(title="Sample")
plt.tight_layout()

plt.savefig(PLOT_FILE, dpi=300)
plt.close()

# Create a second plot without G14 Polymer

plt.figure(figsize=(8, 6))

x = [0, 1, 2]

for sample in sorted(plot_data.keys()):
    if sample != "G14 polymer":
        plt.errorbar(
            x,
            plot_data[sample],
            yerr=plot_unc[sample],
            fmt='o-',
            capsize=6,
            linewidth=2,
            markersize=6,
            label=sample
        )

plt.xticks(x, x_labels)

plt.ylabel("ϕ (°)")
plt.xlabel("ϕ calculation method")
plt.title("Faraday rotation (ϕ) for different samples (without G14 Polymer)")

plt.grid(True, linestyle='--', alpha=0.6)
plt.legend(title="Sample")
plt.tight_layout()

plt.savefig(PLOT_2_FILE, dpi=300)
plt.close()

# ============================================================
# INSERT PLOT INTO EXCEL
# ============================================================

img1 = Image(PLOT_FILE)

img1.anchor = "N2"

ws.add_image(img1)

# Insert second plot

img2 = Image(PLOT_2_FILE)

img2.anchor = "AZ2"

ws.add_image(img2)

# ------------------------------------------------------------
# Save workbook
# ------------------------------------------------------------

wb.save(OUTPUT_FILE)

print(f"\nAnalysis complete.")
print(f"Excel file saved as: {OUTPUT_FILE}")
