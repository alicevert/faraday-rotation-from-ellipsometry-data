import os
import re
from collections import defaultdict

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

# ============================================================
# SETTINGS
# ============================================================

DATA_FOLDER = "angles data"
OUTPUT_EXCEL = "angles.xlsx"
PLOT_FILE = "angles.png"
PLOT_2_FILE = "angles_2.0.png"

# ============================================================
# REGEX PATTERNS
# ============================================================

filename_pattern = re.compile(r"(?P<sample>.+?)_.*?(?P<angle>\d+)-deg_(?P<field>B[+\-0])")

p_pattern = re.compile(r"\bP\s+(-?\d+(?:\.\d+)?)")

psi_pattern = re.compile(r"Psi:\s*(-?\d+(?:\.\d+)?)")

angle_pattern = r"angle-of-incidence:\s*([-\d.]+)"


# ============================================================
# DATA STORAGE
# ============================================================

measurements = {}

grouped_by_angle = defaultdict(dict)

# ============================================================
# READ FILES
# ============================================================

for filename in os.listdir(DATA_FOLDER):

    filepath = os.path.join(DATA_FOLDER, filename)

    if not os.path.isfile(filepath):
        continue

    match = filename_pattern.search(filename)

    if not match:
        print(f"Skipping unrecognized filename: {filename}")
        continue

    unedited_sample = match.group("sample")

    sample = (
        unedited_sample
        .replace("-etch", " etched")
        .replace("-", " ")
    )
    field = match.group("field")

    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    # --------------------------------------------------------
    # Angle of incidence
    # --------------------------------------------------------

    angle_match = re.search(angle_pattern, text)

    if angle_match:
        angle = abs(float(angle_match.group(1)))
        if angle == 43.88: # Force edit B+ field orientation angle to match other angles at 43 degrees
            angle = 43.87
    else:
        print(f"No angle found for {filename}")
        angle = None
        break

    # --------------------------------------------------------
    # Extract P values
    # --------------------------------------------------------

    p_values = [float(x) for x in p_pattern.findall(text)]

    if len(p_values) == 0:
        print(f"No P values found in {filename}")
        break

    p_signs = [1 if p > 0 else -1 for p in p_values]

    if len(set(p_signs)) == 1:
        sign_p = p_signs[0]
    else:
        print(f"Inconsistent signs of polarization angle for {filename}")
        sign_P = None
        break

    # --------------------------------------------------------
    # Extract Psi values
    # --------------------------------------------------------

    psi_values = [float(x) for x in psi_pattern.findall(text)]

    if len(psi_values) == 0:
        print(f"No Psi values found in {filename}")
        break

    psi_avg = float(np.mean(psi_values))

    if len(psi_values) > 1:
        uncert_psi = float(np.std(psi_values, ddof=1))
    else:
        print(f"Insufficient measurements to calculate standard deviation for {filename}")
        uncert_psi = 0.0
        break

    # --------------------------------------------------------
    # Store data
    # --------------------------------------------------------

    measurements[(angle, field)] = {
        "sample": sample,
        "angle": angle,
        "field": field,
        "sign_p": sign_p,
        "psi_avg": psi_avg,
        "uncert_psi": uncert_psi,
    }

    grouped_by_angle[angle][field] = measurements[(angle, field)]

# ============================================================
# TABLE 1: MEASUREMENTS VARYING ANGLE
# ============================================================

measurement_rows = []

for key in sorted(measurements.keys()):

    data = measurements[key]

    measurement_rows.append([
        data["sample"],
        data["angle"],
        data["field"],
        data["sign_p"],
        data["psi_avg"],
        "±",
        data["uncert_psi"],
    ])


measurements_df = pd.DataFrame(
    measurement_rows,
    columns=[
        "Sample",
        "Angle (°)",
        "Magnetic flux density (B)",
        "P sign",
        "Psi (°)",
        "",
        " "
    ]
)

# ============================================================
# TABLE 2: FARADAY ROTATION CALCULATIONS
# ============================================================

faraday_rows = []

plot_data = {}
plot_unc = {}

for angle in sorted(grouped_by_angle.keys()):

    angle_data = grouped_by_angle[angle]

    required_fields = {"B+", "B0", "B-"}

    if not required_fields.issubset(angle_data.keys()):
        print(
            f"Missing one or more of B+, B0, B- files for angle {angle}."
        )

    # Check for consistency of sign P across magnetic field conditions

    bplus = angle_data["B+"]
    b0 = angle_data["B0"]
    bminus = angle_data["B-"]

    try:
        sign_p_bplus = bplus["sign_p"]
        sign_p_b0 = b0["sign_p"]
        sign_p_bminus = bminus["sign_p"]

    except KeyError:
        print(f"Invalid P sign for angle {angle}.")
        break

    else:
        if sign_p_bplus != sign_p_b0 or sign_p_bplus != sign_p_bminus or sign_p_b0 != sign_p_bminus:
            print(f"Inconsistent P signs for angle {angle}.")
            break
        elif sign_p_bplus == sign_p_b0 == sign_p_bminus:
            sign_p_b = sign_p_bplus

    phi_bplus = (bplus["psi_avg"] - b0["psi_avg"]) * sign_p_b

    phi_avg = ((bplus["psi_avg"] - bminus["psi_avg"]) / 2) * sign_p_b

    phi_bminus = (b0["psi_avg"] - bminus["psi_avg"]) * sign_p_b

    # --------------------------------------------------------
    # Uncertainty propagation
    # --------------------------------------------------------

    uncert_phi_bplus = (bplus["uncert_psi"] + b0["uncert_psi"])

    uncert_phi_avg = (bplus["uncert_psi"] + bminus["uncert_psi"])

    uncert_phi_bminus = (b0["uncert_psi"] + bminus["uncert_psi"])

    faraday_rows.append([
        angle,
        sign_p_b,

        phi_bplus,
        "±",
        uncert_phi_bplus,

        phi_avg,
        "±",
        uncert_phi_avg,

        phi_bminus,
        "±",
        uncert_phi_bminus,
    ])

    plot_data[angle] = [
        phi_bplus,
        phi_avg,
        phi_bminus
    ]

    plot_unc[angle] = [
        uncert_phi_bplus,
        uncert_phi_avg,
        uncert_phi_bminus
    ]

faraday_df = pd.DataFrame(
    faraday_rows,
    columns=["Angle (°)", "P sign", "Phi B+ (°)",""," ",
             "Phi avg (°)","  ","   ","Phi B- (°)","    ","     "])

# Organize data by method with phi values corresponding to ascending angle

plot_2_data = {
    'Ψ_B+ - Ψ_B0': [],
    '(Ψ_B+ - Ψ_B0)/2': [],
    'Ψ_B0 - Ψ_B-': [],
}

for angle in plot_data:
    plot_2_data['Ψ_B+ - Ψ_B0'].append(plot_data[angle][0])
    plot_2_data['(Ψ_B+ - Ψ_B0)/2'].append(plot_data[angle][1])
    plot_2_data['Ψ_B0 - Ψ_B-'].append(plot_data[angle][2])

plot_2_unc = {
    'Ψ_B+ - Ψ_B0': [],
    '(Ψ_B+ - Ψ_B0)/2': [],
    'Ψ_B0 - Ψ_B-': [],
}

for angle in plot_unc:
    plot_2_unc['Ψ_B+ - Ψ_B0'].append(plot_unc[angle][0])
    plot_2_unc['(Ψ_B+ - Ψ_B0)/2'].append(plot_unc[angle][1])
    plot_2_unc['Ψ_B0 - Ψ_B-'].append(plot_unc[angle][2])

# ============================================================
# CREATE PLOT
# ============================================================

x = [0, 1, 2]

# Plot 1: phi vs method of Faraday rotation calculation

plt.figure(figsize=(8, 6))

x_labels = ["Ψ_B+ - Ψ_B0", "(Ψ_B+ - Ψ_B-)/2", "Ψ_B0 - Ψ_B-"]

for angle in sorted(plot_data.keys()):

    plt.errorbar(
        x,
        plot_data[angle],
        yerr=plot_unc[angle],
        fmt='o-',
        capsize=6,
        linewidth=2,
        markersize=6,
        label=angle
    )

plt.xticks(x, x_labels)

plt.ylabel("ϕ (°)")
plt.xlabel("ϕ calculation method")
plt.title("Plot 1: Faraday rotation (ϕ) for different laser angles")

plt.grid(True)
plt.legend(title="Angle")

plt.tight_layout()
plt.savefig(PLOT_FILE, dpi=300)
plt.close()

# Plot 2: phi vs angle-of-incidence

plt.figure(figsize=(8, 6))

x_2 = [key for key in sorted(plot_data.keys())]

for B in plot_2_data:
    plt.errorbar(
        x_2,
        y=plot_2_data[B],
        yerr=plot_2_unc[B],
        fmt='o-',
        capsize=6,
        linewidth=2,
        markersize=6,
        label=B
    )

plt.ylabel("ϕ (°)")
plt.xlabel("Angle-of-incidence, relative to normal (°)")
plt.title("Plot 2: Faraday rotation (ϕ) for different laser angles")

plt.grid(True)
plt.legend(title="ϕ calculation method")

plt.tight_layout()
plt.savefig(PLOT_2_FILE, dpi=300)
plt.close()

# ============================================================
# WRITE TO EXCEL
# ============================================================

with pd.ExcelWriter(
    OUTPUT_EXCEL,
    engine="openpyxl"
) as writer:

    sheet_name = "Analysis"

    # --------------------------------------------------------
    # Measurements table
    # --------------------------------------------------------

    measurements_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=2,
        index=False
    )

    # --------------------------------------------------------
    # Faraday table
    # --------------------------------------------------------

    faraday_start_row = (
        len(measurements_df)
        + 8
    )

    faraday_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=faraday_start_row,
        index=False
    )


# ============================================================
# FORMAT EXCEL
# ============================================================

wb = load_workbook(OUTPUT_EXCEL)
ws = wb["Analysis"]

# ------------------------------------------------------------
# Title 1
# ------------------------------------------------------------

ws["A1"] = "Measurements varying angle"
ws["A1"].font = Font(
    bold=True,
    size=14
)

# Merge Psi header
ws.merge_cells("E2:G2")
ws["E2"] = "Psi (°)"
ws["E2"].alignment = Alignment(horizontal="center")
ws["E2"].font = Font(bold=True)

# ------------------------------------------------------------
# Title 2
# ------------------------------------------------------------

faraday_header_row = faraday_start_row + 1

ws.cell(
    row=faraday_start_row,
    column=1
).value = "Faraday rotation calculation varying angle"

ws.cell(
    row=faraday_start_row,
    column=1
).font = Font(
    bold=True,
    size=14
)

# Merge Phi headers
ws.merge_cells(
    start_row=faraday_header_row,
    start_column=3,
    end_row=faraday_header_row,
    end_column=5
)

ws.merge_cells(
    start_row=faraday_header_row,
    start_column=6,
    end_row=faraday_header_row,
    end_column=8
)

ws.merge_cells(
    start_row=faraday_header_row,
    start_column=9,
    end_row=faraday_header_row,
    end_column=11
)

ws.cell(
    row=faraday_header_row,
    column=3
).value = "Phi B+ (°)"

ws.cell(
    row=faraday_header_row,
    column=6
).value = "Phi avg (°)"

ws.cell(
    row=faraday_header_row,
    column=9
).value = "Phi B- (°)"

for col in [3, 6, 9]:
    ws.cell(
        row=faraday_header_row,
        column=col
    ).alignment = Alignment(horizontal="center")
    ws.cell(
        row=faraday_header_row,
        column=col
    ).font = Font(bold=True)

# ------------------------------------------------------------
# Column widths
# ------------------------------------------------------------

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            max_length = max(
                max_length,
                len(str(cell.value))
            )
        except:
            pass

    ws.column_dimensions[column_letter].width = max_length + 2

# ------------------------------------------------------------
# Insert plot
# ------------------------------------------------------------

plot_row = (faraday_start_row + len(faraday_df) + 8)

img = Image(PLOT_FILE)
img.width = 700
img.height = 500

ws.add_image(img,f"A{plot_row}")

# Insert second plot

img = Image(PLOT_2_FILE)
img.width = 700
img.height = 500

ws.add_image(img,f"F{plot_row}")

# ------------------------------------------------------------
# Save workbook
# ------------------------------------------------------------

wb.save(OUTPUT_EXCEL)

print(f"\nAnalysis complete.")
print(f"Excel file saved as: {OUTPUT_EXCEL}")